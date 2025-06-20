import pandas as pd
import numpy as np
from typing import List, Dict, Optional, Tuple
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from datetime import datetime

from data.models import Store, Supplier, Location, Vehicle, Order, PalletType, Route, OptimizationResult


class ExcelHandler:
    def __init__(self, input_directory: str = "data/input", output_directory: str = "data/output"):
        self.input_dir = Path(input_directory)
        self.output_dir = Path(output_directory)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def load_stores(self, filename: str = "store_locations.xlsx") -> List[Store]:
        file_path = self.input_dir / filename
        if not file_path.exists():
            raise FileNotFoundError(f"Store data file not found: {file_path}")
        
        df = pd.read_excel(file_path)
        stores = []
        
        for _, row in df.iterrows():
            # Handle both 'name' and 'store_name' column formats
            store_name = row.get('name', row.get('store_name', 'Unknown Store'))
            
            location = Location(
                name=store_name,
                address=row['address'],
                latitude=row['latitude'],
                longitude=row['longitude'],
                city=row['city'],
                state=row['state'],
                zip_code=str(row['zip_code']),
                contact_info=row.get('contact_info')
            )
            
            delivery_start = None
            delivery_end = None
            if 'delivery_window_start' in row and pd.notna(row['delivery_window_start']):
                delivery_start = pd.to_datetime(row['delivery_window_start'])
            if 'delivery_window_end' in row and pd.notna(row['delivery_window_end']):
                delivery_end = pd.to_datetime(row['delivery_window_end'])
            
            # Handle priority mapping from text to numbers
            priority_map = {'High': 1, 'Medium': 2, 'Low': 3, 'high': 1, 'medium': 2, 'low': 3}
            priority_value = row.get('priority', 1)
            if isinstance(priority_value, str):
                priority_value = priority_map.get(priority_value, 1)
            else:
                priority_value = int(priority_value)
            
            store = Store(
                id=str(row['store_id']),
                name=store_name,
                location=location,
                demand_pallets=int(row['demand_pallets']),
                delivery_window_start=delivery_start,
                delivery_window_end=delivery_end,
                priority=priority_value
            )
            stores.append(store)
        
        return stores
    
    def load_suppliers(self, filename: str = "supplier_data.xlsx") -> List[Supplier]:
        file_path = self.input_dir / filename
        if not file_path.exists():
            raise FileNotFoundError(f"Supplier data file not found: {file_path}")
        
        df = pd.read_excel(file_path)
        suppliers = []
        
        for _, row in df.iterrows():
            # Handle both 'name' and 'supplier_name' column formats
            supplier_name = row.get('name', row.get('supplier_name', 'Unknown Supplier'))
            
            location = Location(
                name=supplier_name,
                address=row['address'],
                latitude=row['latitude'],
                longitude=row['longitude'],
                city=row['city'],
                state=row['state'],
                zip_code=str(row['zip_code']),
                contact_info=row.get('contact_info')
            )
            
            pallet_types = []
            if 'pallet_types' in row and pd.notna(row['pallet_types']):
                type_str = str(row['pallet_types']).lower()
                if 'standard' in type_str:
                    pallet_types.append(PalletType.STANDARD)
                if 'euro' in type_str:
                    pallet_types.append(PalletType.EURO)
                if 'custom' in type_str:
                    pallet_types.append(PalletType.CUSTOM)
            
            if not pallet_types:
                pallet_types = [PalletType.STANDARD]
            
            supplier = Supplier(
                id=str(row['supplier_id']),
                name=supplier_name,
                location=location,
                available_pallets=int(row['available_pallets']),
                cost_per_pallet=float(row['cost_per_pallet']),
                lead_time_days=int(row.get('lead_time_days', 1)),
                capacity_per_day=int(row.get('capacity_per_day', 100)),
                reliability_score=float(row.get('reliability_score', 1.0)),
                pallet_types=pallet_types
            )
            suppliers.append(supplier)
        
        return suppliers
    
    def load_historical_orders(self, filename: str = "historical_orders.xlsx") -> List[Order]:
        file_path = self.input_dir / filename
        if not file_path.exists():
            return []
        
        df = pd.read_excel(file_path)
        orders = []
        
        for _, row in df.iterrows():
            pallet_type = PalletType.STANDARD
            if 'pallet_type' in row and pd.notna(row['pallet_type']):
                pallet_type = PalletType(row['pallet_type'].lower())
            
            # Handle quantity column - could be 'quantity' or 'pallets_ordered'
            quantity = row.get('quantity', row.get('pallets_ordered', 0))
            
            # Handle requested_date - could be 'requested_date' or 'date'
            date_field = row.get('requested_date', row.get('date'))
            if date_field:
                requested_date = pd.to_datetime(date_field)
            else:
                requested_date = pd.Timestamp.now()
            
            # Handle priority - could be text or number
            priority_map = {'High': 1, 'Medium': 2, 'Low': 3}
            priority_value = row.get('priority', row.get('order_priority', 1))
            if isinstance(priority_value, str):
                priority_value = priority_map.get(priority_value, 1)
            else:
                priority_value = int(priority_value) if priority_value else 1
            
            order = Order(
                id=str(row['order_id']),
                store_id=str(row['store_id']),
                supplier_id=str(row['supplier_id']),
                quantity=int(quantity),
                pallet_type=pallet_type,
                requested_date=requested_date,
                priority=priority_value,
                special_instructions=row.get('special_instructions', '')
            )
            orders.append(order)
        
        return orders
    
    def load_toll_rates(self, filename: str = "toll_rates.xlsx") -> Dict[Tuple[str, str], float]:
        file_path = self.input_dir / filename
        toll_rates = {}
        
        if file_path.exists():
            df = pd.read_excel(file_path)
            for _, row in df.iterrows():
                # Handle different toll rate file formats
                if 'from_location' in row and 'to_location' in row:
                    # Standard format
                    key = (str(row['from_location']), str(row['to_location']))
                    toll_rates[key] = float(row['rate_per_mile'])
                elif 'route_segment' in row:
                    # Frito-Lay format - parse route segment
                    route_segment = str(row['route_segment'])
                    # Extract rate - could be 'toll_rate_per_mile' or 'rate_per_mile'
                    rate = row.get('toll_rate_per_mile', row.get('rate_per_mile', 0.0))
                    
                    # Try to extract from/to from route segment
                    if 'to' in route_segment.lower():
                        parts = route_segment.split(' to ')
                        if len(parts) == 2:
                            from_loc = parts[0].strip()
                            to_loc = parts[1].strip()
                            key = (from_loc, to_loc)
                            toll_rates[key] = float(rate)
                    else:
                        # Use route segment as a single key
                        key = (route_segment, route_segment)
                        toll_rates[key] = float(rate)
        
        return toll_rates
    
    def save_optimization_results(self, result: OptimizationResult, filename: str = None):
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"optimization_results_{timestamp}.xlsx"
        
        file_path = self.output_dir / filename
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Metric': ['Total Cost', 'Total Distance', 'Total Time', 'Number of Routes', 
                          'Utilization Rate', 'Solver Status', 'Solve Time', 'Objective Value'],
                'Value': [f"${result.total_cost:,.2f}", f"{result.total_distance:.1f} miles",
                         f"{result.total_time:.1f} hours", len(result.routes),
                         f"{result.utilization_rate:.1%}", result.solver_status,
                         f"{result.solve_time:.2f} seconds", f"{result.objective_value:,.2f}"]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Routes sheet
            routes_data = []
            for route in result.routes:
                routes_data.append({
                    'Route ID': route.id,
                    'Vehicle ID': route.vehicle_id,
                    'Stops': ' -> '.join(route.stops),
                    'Distance (miles)': route.total_distance,
                    'Time (hours)': route.total_time,
                    'Cost': route.total_cost,
                    'Pallets': route.pallets_delivered,
                    'Status': route.status.value
                })
            
            if routes_data:
                routes_df = pd.DataFrame(routes_data)
                routes_df.to_excel(writer, sheet_name='Routes', index=False)
            
            # Format the workbook
            self._format_workbook(writer.book)
    
    def _format_workbook(self, workbook):
        # Apply formatting to all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_template_files(self):
        # Create store template - use 'name' column to match actual data files
        store_template = pd.DataFrame({
            'store_id': ['STORE_001', 'STORE_002'],
            'name': ['Store A', 'Store B'],  # Fixed: changed from 'store_name' to 'name'
            'address': ['123 Main St', '456 Oak Ave'],
            'city': ['Chicago', 'Milwaukee'],
            'state': ['IL', 'WI'],
            'zip_code': ['60601', '53202'],
            'latitude': [41.8781, 43.0389],
            'longitude': [-87.6298, -87.9065],
            'demand_pallets': [25, 15],
            'priority': [1, 2],
            'contact_info': ['manager@storea.com', 'contact@storeb.com']
        })
        
        # Create supplier template - use 'name' column to match actual data files
        supplier_template = pd.DataFrame({
            'supplier_id': ['SUP_001', 'SUP_002'],
            'name': ['Supplier Alpha', 'Supplier Beta'],  # Fixed: changed from 'supplier_name' to 'name'
            'address': ['789 Industrial Blvd', '321 Commerce Dr'],
            'city': ['Chicago', 'Milwaukee'],
            'state': ['IL', 'WI'],
            'zip_code': ['60610', '53210'],
            'latitude': [41.8902, 43.0642],
            'longitude': [-87.6511, -87.9073],
            'available_pallets': [100, 75],
            'cost_per_pallet': [45.50, 42.00],
            'lead_time_days': [1, 2],
            'capacity_per_day': [50, 40],
            'reliability_score': [0.95, 0.98],
            'pallet_types': ['standard,euro', 'standard']
        })
        
        # Save templates
        template_dir = Path("excel_templates")
        template_dir.mkdir(exist_ok=True)
        
        store_template.to_excel(template_dir / "store_input_template.xlsx", index=False)
        supplier_template.to_excel(template_dir / "supplier_input_template.xlsx", index=False)